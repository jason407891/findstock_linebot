import requests
import openpyxl
import json
import time


def getbreak(price_breaks_data,qty):
    goods_num = 1
    if price_breaks_data:
        price=price_breaks_data[0]['Price']
        for price_break in price_breaks_data:
            if qty <= price_break['Quantity']:
                break
            goods_num = price_break['Quantity']
            price=price_break['Price']
        return [goods_num,price]
    else:
        return["NA","NA"]

def getdata(pn):
    # Swagger API 文件中的路徑和參數
    url = 'https://api.mouser.com/api/v1/search/partnumber?apiKey=8b1390e5-5fbc-4169-9c88-ac16c0599220'
    api_key_order = '19c5e8f6-6454-4387-bbbd-fbfa3b4d1434'

    # 設定請求的 headers，如果需要 API 金鑰
    headers = {
        'Authorization': f'Bearer {api_key_order}',
        'Content-Type': 'application/json'
    }

    # 設定請求的資料，這裡使用 JSON 格式
    data = {
        "SearchByPartRequest": {
            "mouserPartNumber": pn,
            "partSearchOptions": "string"
        }
    }

    # 發送 POST 請求
    response = requests.post(url,headers=headers,json=data)

    # 解析回應內容
    response_content = response.json()

    # 檢查回應是否為空
    if response.status_code == 200:
        # 解析回應內容
        response_content = response.json()

        # 檢查回應是否包含 SearchResults
        if response_content and 'SearchResults' in response_content:
            # 取得 Parts
            parts = response_content['SearchResults']['Parts']
        
        # 逐個處理每個部分
            for part in parts:
                availability = part.get('Availability')
                manufacturer = part.get('Manufacturer')
                manufacturer_part_number = part.get('ManufacturerPartNumber')
                price_breaks = part.get('PriceBreaks', [])
                
                all_parts_info = {}
                # 顯示資訊
                part_info = {
                    'Availability': availability,
                    'Manufacturer': manufacturer,
                    'ManufacturerPartNumber': manufacturer_part_number,
                    'PriceBreaks': price_breaks
                }
                
                all_parts_info[manufacturer_part_number] = part_info
                
                return all_parts_info
        else:
            return {"nodata"}
    else:
        return {"nodata"}


wb = openpyxl.load_workbook('input.xlsx')
ws = wb['工作表1']
items = []
qtys=[]
for row in ws.iter_rows(min_row=2, values_only=True):  # 從第二行開始讀取
    if row[0]:  # 如果該行的第一個欄位不為空
        items.append(row[0])
        qtys.append(row[1])


# 建立新的 Excel 檔案
output_wb = openpyxl.Workbook()
output_ws = output_wb.active
output_ws.append(["搜尋編號","庫存","製造商","產品編號","數量級距","價格(USD)"])
qtyposition=0
# 在新的 Excel 檔案中寫入資料
for item in items:
    #qty對應的位子
    qtyvalue=qtys[qtyposition]
    print(item)
    time.sleep(1)
    result = getdata(item)
    if result != {"nodata"} and result:
        for part_number, part_info in result.items():
            row = [part_number, part_info['Availability'], part_info['Manufacturer'], part_info['ManufacturerPartNumber']]
            
            price_breaks = part_info['PriceBreaks']
            print(price_breaks)
            breakprice=getbreak(price_breaks,qtyvalue)
            row.append(breakprice[0])
            row.append(breakprice[1])

            output_ws.append(row)
    else:
        output_ws.append([item,"NA"])
    qtyposition+=1

# 儲存新的 Excel 檔案
output_wb.save('output.xlsx')



